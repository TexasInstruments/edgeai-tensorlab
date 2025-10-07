import os
import glob
import argparse
import pandas as pd
import subprocess

def parse_args():
    parser = argparse.ArgumentParser(description="Generate Benchmark Report")
    parser.add_argument("--report_path", required=True, help="Path to the report.csv")
    parser.add_argument("--ref_report_path", required=True, help="Path to the reference report.csv")
    args = parser.parse_args()
    return args

def generate_benchmark_report(report, ref_report):
    """
    Compare the specified metric column in report with best_metric in ref_report.
    Group as improved, degraded, Requires Review, disabled, enabled, inactive, or Same.
    """
    report_metric_col = report.columns[8]

    results = {
        "Degraded": [],
        "Improved": [],
        "Requires Review": [],
        "Disabled": [],
        "Enabled": [],
        "Same": [],
        "Inactive": [],
    }

    # Merge on 'model_id'
    merged = pd.merge(report, ref_report, on="model_id", suffixes=('_new', '_ref'))

    def is_numeric(val):
        try:
            float(val)
            return True
        except (ValueError, TypeError):
            return False

    for _, row in merged.iterrows():
        new_val = row[f"{report_metric_col}_new"]
        ref_val = row[f"{report_metric_col}_ref"]
        model_id = row["model_id"]

        new_is_num = is_numeric(new_val)
        ref_is_num = is_numeric(ref_val)

        if new_is_num and ref_is_num:
            new_val = float(new_val)
            ref_val = float(ref_val)
            if new_val > ref_val:
                results["Improved"].append(model_id)
            elif new_val < ref_val * 0.998:
                results["Degraded"].append(model_id)
            elif new_val == ref_val:
                results["Same"].append(model_id)
            else:
                results["Requires Review"].append(model_id)
        elif not new_is_num and ref_is_num:
            results["Disabled"].append(model_id)
        elif new_is_num and not ref_is_num:
            results["Enabled"].append(model_id)
        else:  # both not numeric
            results["Inactive"].append(model_id)

    return results

def save_results_to_excel(results, output_path="report_comparison.xlsx"):
    """
    Save all groups in a single worksheet, each group as a colored banner row (merged and centered), followed by full rows from report with a diff column, and a gap row between groups.
    Banner and diff column are colored.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment

    group_colors = {
        "Improved": "00FF00",      # green
        "Degraded": "FF0000",      # red
        "Requires Review": "FFFF00", # yellow
        "Disabled": "FFC7CE",      # light red
        "Enabled": "BDD7EE",
        "Same": "B7E1CD",
        "Inactive": "D9D9D9"       # grey
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Benchmark Comparison"

    # Get report DataFrame and ref_sha from caller's frame
    from inspect import currentframe
    frame = currentframe().f_back
    report = frame.f_locals.get('report')
    ref_report = frame.f_locals.get('ref_report')
    ref_sha = frame.f_locals.get('ref_sha')
    report_metric_col = report.columns[8]
    # ref_metric_col = "best_metric"
    ref_metric_col = ref_report.columns[8]
    num_cols = len(report.columns) + 1

    row_idx = ws.max_row
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=num_cols)
    cell = ws.cell(row=row_idx, column=1)
    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    # Add gap row under the top row
    ws.append([""] * num_cols)

    for group, model_ids in results.items():
        # Banner row (merged and centered)
        ws.append([group] + [""] * (num_cols - 1))
        row_idx = ws.max_row  # get the row just appended
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=num_cols)
        fill = PatternFill(start_color=group_colors.get(group, "FFFFFF"), end_color=group_colors.get(group, "FFFFFF"), fill_type="solid")
        cell = ws.cell(row=row_idx, column=1)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

        group_rows = report[report["model_id"].isin(model_ids)]
        if not group_rows.empty:
            # Header row (report columns + diff)
            header = list(report.columns) + ["diff"]
            ws.append(header)
            # Write data rows
            for _, row in group_rows.iterrows():
                model_id = row["model_id"]
                ref_row = ref_report[ref_report["model_id"] == model_id]
                if not ref_row.empty:
                    ref_val = ref_row.iloc[0][ref_metric_col]
                else:
                    ref_val = None
                new_val = row[report_metric_col]
                try:
                    diff = float(new_val) - float(ref_val) if ref_val is not None and pd.notnull(new_val) and pd.notnull(ref_val) else ""
                except Exception:
                    diff = ""
                values = [row[col] for col in report.columns] + [diff]
                ws.append(values)
                # Color only the diff column
                diff_col_idx = len(report.columns) + 1
                ws.cell(row=ws.max_row, column=diff_col_idx).fill = fill
        # Add gap row
        ws.append([""] * num_cols)
    wb.save(output_path)


def main(report_path, ref_report_path):
    report = pd.read_csv(report_path)
    modelartifacts_path = os.path.dirname(report_path)

    ref_report = pd.read_csv(ref_report_path)

    results = generate_benchmark_report(report, ref_report)
    print("Improved:", results["Improved"])
    print("Degraded:", results["Degraded"])
    print("Requires Review:", results["Requires Review"])
    print("Disabled:", results["Disabled"])
    print("Enabled:", results["Enabled"])
    print("Inactive:", results["Inactive"])
    print("Same:", results["Same"])

    # Save results to Excel
    save_results_to_excel(results, os.path.join(modelartifacts_path, "report_comparison.xlsx"))


if __name__ == "__main__":
    args = parse_args()
    main(args.report_path, args.ref_report_path)